from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / "Book1.xlsx"
TARGET = ROOT / "src" / "utils" / "seed.book1.assets.ts"


def display_name(name: str, variant: str | None) -> str:
    if variant and variant.startswith("AIM_"):
        return variant.replace("_", "-")
    return name


def main() -> None:
    workbook = load_workbook(BOOK, read_only=True, data_only=True)
    if "Asset IDs" not in workbook.sheetnames:
        raise SystemExit(f'Missing "Asset IDs" sheet in {BOOK}')
    sheet = workbook["Asset IDs"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    workbook.close()

    lines = [
        'import type { Asset } from "@/lib/models";',
        "",
        "export function buildBook1Assets(): Asset[] {",
        "  return [",
    ]

    for row in rows:
        if not row or not row[0]:
            continue
        asset_id, category, name, specs, serial, _upgrades, variant = row
        display = display_name(str(name), str(variant) if variant else None)
        lines.extend(
            [
                "    {",
                f'      id: "{asset_id}",',
                f"      name: {json.dumps(display)},",
                f'      category: "{category}",',
                "      assignedTo: null,",
                '      department: "IT",',
                '      status: "Active",',
            ]
        )
        if specs:
            lines.append(f"      specifications: {json.dumps(str(specs))},")
        if serial:
            lines.append(f'      serial: {json.dumps(str(serial))},')
        lines.append("    },")

    lines.extend(["  ];", "}"])
    TARGET.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {len([row for row in rows if row and row[0]])} assets to {TARGET}")


if __name__ == "__main__":
    main()
