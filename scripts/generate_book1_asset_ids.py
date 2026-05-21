from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / "Book1.xlsx"

INVENTORY = [
    {
        "label": "Aim_A",
        "asset_type": "PC BUILT",
        "count": 5,
        "specs": [
            "Ram-64GB",
            "SSD-1TB",
            "HDD-/",
            "Processor-I9-12th gen",
            "GPU-8 GB RTX 3060 TI",
        ],
        "upgrades": "Up graded from 32 to 64 gb ram",
        "codes": [],
    },
    {
        "label": "Aim_B",
        "asset_type": "PC BUILT",
        "count": 2,
        "specs": [
            "Ram-32GB",
            "SSD-500GB",
            "HDD-1TB",
            "Processor- AMD RYZEN 5",
            "GPU-8 GB RTX 3060 TI",
        ],
        "upgrades": "",
        "codes": [],
    },
    {
        "label": "Aim_C",
        "asset_type": "PC BUILT",
        "count": 32,
        "specs": [
            "Ram-16GB",
            "SSD-1TB",
            "HDD-/",
            "Processor- AMD RYZEN 5",
            "GPU-8 GB AMD RADION RX 6600",
        ],
        "upgrades": "10PCS  Up graded from 16 to 32 gb ram",
        "codes": ["C01", "C14", "C07", "C19", "C29", "C13", "C20", "C04", "C06", "C15"],
    },
    {
        "label": "Aim_D",
        "asset_type": "PC BUILT",
        "count": 2,
        "specs": [
            "Ram-48GB",
            "SSD-1TB",
            "HDD-/",
            "Processor-AMD RYZEN 7",
            "GPU-8GB RADION RX 580 2048 SP",
        ],
        "upgrades": "",
        "codes": [],
    },
    {
        "label": "Aim_E",
        "asset_type": "PC BUILT",
        "count": 2,
        "specs": [
            "Ram-48GB",
            "SSD-1TB",
            "HDD-/",
            "Processor-AMD RYZEN 7",
            "GPU-4GB RTX 730",
        ],
        "upgrades": "",
        "codes": [],
    },
    {
        "label": "Aim_F",
        "asset_type": "PC BUILT",
        "count": 2,
        "specs": [
            "Ram-16GB",
            "SSD-500GB",
            "HDD-/",
            "Processor- I5",
            "GPU-4 GB RTX 730",
        ],
        "upgrades": "",
        "codes": [],
    },
    {
        "label": "LAP-01",
        "asset_type": "LAPTOP",
        "count": 1,
        "specs": [
            "Ram-32GB",
            "SSD-1TB",
            "HDD-/",
            "Processor-I7 12th gen",
            "GPU-8 GB RTX 3070 TI",
        ],
        "upgrades": "",
        "codes": [],
    },
    {
        "label": "MON_B",
        "asset_type": "Lenovo monitors",
        "count": 1,
        "specs": [],
        "upgrades": "",
        "codes": [],
    },
    {
        "label": "MON_BT",
        "asset_type": "Think center Monitors",
        "count": 1,
        "specs": [],
        "upgrades": "",
        "codes": [],
    },
    {
        "label": "MON_BA",
        "asset_type": "AOC Monitors",
        "count": 1,
        "specs": [],
        "upgrades": "",
        "codes": [],
    },
    {
        "label": "MON_S",
        "asset_type": "Samsung monitors",
        "count": 1,
        "specs": [],
        "upgrades": "",
        "codes": [],
    },
    {
        "label": "M_01",
        "asset_type": "MOUSE",
        "count": 1,
        "specs": [],
        "upgrades": "",
        "codes": [],
    },
    {
        "label": "K_01",
        "asset_type": "KEYBOARD",
        "count": 1,
        "specs": [],
        "upgrades": "",
        "codes": [],
    },
    {
        "label": "H_01",
        "asset_type": "HEADSET",
        "count": 1,
        "specs": [],
        "upgrades": "",
        "codes": [],
    },
]


MONITOR_VARIANTS = {
    "MON_B": ("MON_LENOVO", "Lenovo monitor"),
    "MON_BT": ("MON_THINKCENTRE", "ThinkCentre monitor"),
    "MON_BA": ("MON_AOC", "AOC monitor"),
    "MON_S": ("MON_SAMSUNG", "Samsung monitor"),
}

ACCESSORY_VARIANTS = {
    "M_01": ("ACC_MOUSE", "Mouse"),
    "K_01": ("ACC_KEYBOARD", "Keyboard"),
    "H_01": ("ACC_HEADSET", "Headset"),
}


def build_assignments() -> list[tuple[str, str, str, str, str | None, str, str]]:
    counters = {"PC": 0, "LAP": 0, "MON": 0, "ACC": 0}
    assignments: list[tuple[str, str, str, str, str | None, str, str]] = []

    def next_id(kind: str) -> str:
        counters[kind] += 1
        return f"{kind}-{counters[kind]:04d}"

    for group in INVENTORY:
        label = group["label"]
        asset_type = group["asset_type"]
        specs = "; ".join(group["specs"])
        upgrades = group["upgrades"]

        if label.startswith("Aim_") or asset_type == "PC BUILT":
            for index in range(group["count"]):
                asset_id = next_id("PC")
                serial = group["codes"][index] if index < len(group["codes"]) else None
                assignments.append(
                    (
                        asset_id,
                        "Desktop",
                        f"AIM-{label.split('_')[1]}",
                        specs,
                        serial,
                        upgrades,
                        f"AIM_{label.split('_')[1]}",
                    )
                )
            continue

        if label.startswith("LAP") or asset_type == "LAPTOP":
            assignments.append(
                (next_id("LAP"), "Laptop", "Laptop workstation", specs, None, upgrades, "LAP_01")
            )
            continue

        if label.startswith("MON_"):
            variant, display_name = MONITOR_VARIANTS[label]
            assignments.append((next_id("MON"), "Monitor", display_name, specs, None, upgrades, variant))
            continue

        variant, display_name = ACCESSORY_VARIANTS[label]
        assignments.append((next_id("ACC"), "Accessory", display_name, specs, None, upgrades, variant))

    return assignments


def group_asset_ids(group: dict, assignments: list[tuple]) -> list[str]:
    label = group["label"]
    if label.startswith("Aim_"):
        variant = f"AIM_{label.split('_')[1]}"
        return [row[0] for row in assignments if row[6] == variant]
    if label.startswith("LAP"):
        return [row[0] for row in assignments if row[6] == "LAP_01"]
    if label.startswith("MON_"):
        variant, _ = MONITOR_VARIANTS[label]
        return [row[0] for row in assignments if row[6] == variant]
    variant, _ = ACCESSORY_VARIANTS[label]
    return [row[0] for row in assignments if row[6] == variant]


def write_inventory_sheet(ws, assignments: list[tuple]) -> None:
    ws.append(["Asset ID", "Asset Type", "Count", "specification", "UPGRADES"])
    for group in INVENTORY:
        group_ids = group_asset_ids(group, assignments)
        primary_id = group_ids[0] if len(group_ids) == 1 else f"{group_ids[0]}..{group_ids[-1]}"
        ws.append([primary_id, "", group["count"], group["specs"][0] if group["specs"] else "", group["upgrades"]])

        if group["specs"]:
            for spec in group["specs"][1:]:
                ws.append(["", "", "", spec, ""])

        if group["asset_type"] in {"PC BUILT", "LAPTOP"}:
            ws.cell(ws.max_row - len(group["specs"]) + 1, 2).value = group["asset_type"]
        elif not group["label"].startswith("Aim_") and not group["label"].startswith("LAP"):
            ws.cell(ws.max_row - max(len(group["specs"]) - 1, 0), 2).value = group["asset_type"]

        if group["codes"]:
            row = ws.max_row - len(group["specs"]) + 3
            for index, code in enumerate(group["codes"]):
                ws.cell(row, 6 + index).value = code

        ws.append([])


def write_asset_ids_sheet(wb, assignments: list[tuple]) -> None:
    if "Asset IDs" in wb.sheetnames:
        del wb["Asset IDs"]
    sheet = wb.create_sheet("Asset IDs")
    sheet.append(["Asset ID", "Category", "Name", "Specifications", "Serial / Unit", "Upgrades", "Variant"])
    for row in assignments:
        sheet.append(list(row))


def main() -> None:
    assignments = build_assignments()
    wb = Workbook()
    write_inventory_sheet(wb.active, assignments)
    wb.active.title = "Inventory"
    write_asset_ids_sheet(wb, assignments)
    wb.save(BOOK)

    print(f"Generated {len(assignments)} asset IDs in {BOOK}")
    for asset_id, category, name, *_ in assignments:
        print(f"{asset_id} | {category} | {name}")


if __name__ == "__main__":
    main()
