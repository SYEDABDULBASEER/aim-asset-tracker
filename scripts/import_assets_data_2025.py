#!/usr/bin/env python3
"""
Import "PCs and Specifications" sheet from Assets Data 2025.xlsx.

Column mapping (sheet → Asset model):
  S No         → serial
  Name         → name (id = PC-{Name code}, e.g. PC-A01)
  Accessories  → specifications (Accessories: …)
  Softwares    → specifications (Softwares: …)
  Desk         → location
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from lib.asset_import import HEADERS, build_specifications, parse_sheet  # noqa: E402

WORKBOOK = ROOT / "Assets Data 2025.xlsx"
SHEET_NAME = "PCs and Specifications"
TARGET = ROOT / "src" / "utils" / "seed.assets-data-2025.ts"


def main() -> None:
    if not WORKBOOK.is_file():
        raise SystemExit(f"Workbook not found: {WORKBOOK}")

    wb = load_workbook(WORKBOOK, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f'Missing "{SHEET_NAME}". Sheets: {wb.sheetnames}')

    ws = wb[SHEET_NAME]
    row1 = [ws.cell(1, c).value for c in range(1, 6)]
    if [str(h).strip() if h else "" for h in row1] != list(HEADERS):
        raise SystemExit(f"Unexpected headers in row 1: {row1}\nExpected: {HEADERS}")

    assets = parse_sheet(ws)
    wb.close()

    lines = [
        'import type { Asset } from "@/lib/models";',
        "",
        "/**",
        f' * Sheet "{SHEET_NAME}" from {WORKBOOK.name}',
        " * Columns: S No → serial, Name → name, Desk → location,",
        " * Accessories + Softwares → specifications.",
        " * Regenerate: npm run import:assets-2025",
        " */",
        "export function buildAssetsData2025(): Asset[] {",
        "  return [",
    ]

    for asset in assets:
        specs = build_specifications(asset["accessories"], asset["softwares"])
        lines.extend(
            [
                "    {",
                f'      id: {json.dumps(asset["id"])},',
                f"      name: {json.dumps(asset['name'])},",
                '      category: "Desktop",',
                "      assignedTo: null,",
                "      department: null,",
                '      status: "Active",',
                "      warrantyUntil: null,",
                "      lastServiceAt: null,",
                f"      serial: {json.dumps(asset['serial'] or None)},",
                f"      location: {json.dumps(asset['location'])},",
                "      purchaseDate: null,",
                f"      specifications: {json.dumps(specs or None)},",
                "    },",
            ]
        )

    lines.extend(["  ];", "}"])
    TARGET.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {len(assets)} assets ({', '.join(HEADERS)}) → {TARGET}")


if __name__ == "__main__":
    main()
