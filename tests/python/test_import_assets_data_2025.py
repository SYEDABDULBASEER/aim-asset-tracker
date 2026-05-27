from __future__ import annotations

from openpyxl import load_workbook

from lib.asset_import import (
    HEADERS,
    build_specifications,
    format_sno,
    is_pc_name,
    normalize_id,
    parse_sheet,
)


def test_normalize_id_pc_code():
    assert normalize_id("A01 Laptop") == "PC-A01"


def test_normalize_id_slug_fallback():
    assert normalize_id("Special Device").startswith("PC-")


def test_is_pc_name_rejects_headers():
    assert is_pc_name("RAM") is False
    assert is_pc_name("CATEGORY") is False


def test_is_pc_name_accepts_codes():
    assert is_pc_name("A12") is True


def test_format_sno_float_integer():
    assert format_sno(42.0) == "42"


def test_build_specifications_accessories_only():
    assert build_specifications(["Mouse"], []) == "Accessories: Mouse"


def test_parse_sheet_merged_desk(minimal_pcs_workbook):
    wb = load_workbook(minimal_pcs_workbook, data_only=True)
    ws = wb["PCs and Specifications"]
    assets = parse_sheet(ws)
    wb.close()
    assert len(assets) == 1
    assert assets[0]["id"] == "PC-A01"
    assert assets[0]["location"] == "Desk A"
    assert "Mouse" in assets[0]["accessories"]
    assert "Webcam" in assets[0]["accessories"]


def test_parse_sheet_empty(empty_pcs_workbook):
    wb = load_workbook(empty_pcs_workbook, data_only=True)
    ws = wb["PCs and Specifications"]
    assert parse_sheet(ws) == []
    wb.close()


def test_headers_constant():
    assert HEADERS == ("S No", "Name", "Accessories", "Softwares", "Desk")
