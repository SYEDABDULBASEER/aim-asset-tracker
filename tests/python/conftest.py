from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"


@pytest.fixture
def root_dir() -> Path:
    return ROOT


@pytest.fixture
def minimal_pcs_workbook(tmp_path: Path) -> Path:
    """Small workbook matching HEADERS with one PC and merged desk cell."""
    path = tmp_path / "minimal_pcs.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "PCs and Specifications"
    ws.append(["S No", "Name", "Accessories", "Softwares", "Desk"])
    ws.append([1, "A01", "Mouse", "Windows 11", "Desk A"])
    ws.append([None, None, "Webcam", None, None])
    ws.merge_cells("E2:E3")
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def empty_pcs_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "empty_pcs.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "PCs and Specifications"
    ws.append(["S No", "Name", "Accessories", "Softwares", "Desk"])
    wb.save(path)
    wb.close()
    return path
